"Module for creating, saving exporting and plotting data from and to a database."
import os
import sqlite3
from sqlite3 import Error
import datetime as dt
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from matplotlib import pyplot


class SimsimsAnalytics:
    """
    Handles the data logging, storage, and analysis of Simsims simulation
    with metrics such as products, workers and food.
    The SimSimsAnalytics class creates a DataBase and stores the simulation 
    data of the simsims simulation. It also saves data to excel and plot
    a graph over the resource amounts in the simulation.

    Attributes:
        db_file (str): Path and name of the SQLite database file.
        table_columns (list[str]): Columns to track in the simulation.
        
    Methods:
        create_table(): Creates a table in the database.
        drop_table(): Deletes the table from the Database.
        __str__(): Returns the recent row.
        get_rows(): Retrieves all rows from the table.
        add_step(data): Adds a record to the table in the DB.
        to_excel(filename): Exports table data to an Excel file.
        to_figure(filename): Generates a plot from the database.
    """
    def __init__(self, db_file, table_columns = ['Worker', 'Product', 'Food']):
        self._sim_id: int
        self.__table_name = 'simulations'
        self.__folder_name = 'Loggs'
        self.__table_columns = table_columns

        if '.db' not in db_file:
            db_file += '.db'

        try:
            self.__db = self._create_connection(self.__getpath(self.__folder_name) + '\\' + db_file)
        except FileNotFoundError:
            print("Reatempting without the specific folder path")
            self.__db = self._create_connection(db_file)
        self.__db_c = self.__db.cursor() # Database cursor

    def __getpath(self, filename):
        return os.path.join(os.path.join(os.path.dirname(__file__)), filename)

    def _create_connection(self, db_file = ''):
        conn = None
        try:
            conn = sqlite3.connect(db_file)
            return conn
        except Error as e:
            print(e, " with " + db_file)
            raise FileNotFoundError from e



    def create_table(self, table_name = 'simulations'):
        "Create a table in the database."
        if table_name == '':
            using_table_name = self.__table_name
        else:
            using_table_name = table_name

        sql_query = f"""
            CREATE TABLE IF NOT EXISTS {using_table_name}(
            SIM_ID INTEGER PRIMARY KEY AUTOINCREMENT,
            DATE DATETIME,
            """

        for i, column in enumerate(self.__table_columns):
            if i == len(self.__table_columns) - 1:  # If it's the last column
                sql_query += f"{column} INTEGER DEFAULT 0\n);"
            else:
                sql_query += f"{column} INTEGER DEFAULT 0,\n"

        self.__db_c.execute(sql_query)
        self.__db.commit()

    def drop_table(self, table_name = ''):
        "Drop the table in the database."
        if table_name == '':
            using_table_name = self.__table_name
        else:
            using_table_name = table_name
        sql_query = f"""
            DROP TABLE IF EXISTS {using_table_name};
            """
        self.__db_c.execute(sql_query)
        self.__db.commit()

    def __str__(self, table_name = 'simulations') -> str:
        if not self.__check_table_exists():
            return '' # The table does not exist.

        if table_name == '':
            using_table_name = self.__table_name
        else:
            using_table_name = table_name

        sql_query = f"""
                SELECT * FROM {using_table_name};
                """
        self.__db_c.execute(sql_query)
        string = self.__db_c.fetchall()[-1]  # Get the last row.
        return str(string)

    def __check_table_exists(self):
        # Check if the table exists by querying the system table 'sqlite_master'
        table_exists_query = f"""
            SELECT name FROM sqlite_master WHERE type='table'
            AND name='{self.__table_name}';
        """
        self.__db_c.execute(table_exists_query)
        # Returns None if table doesn't exist, True if it does.
        return False if self.__db_c.fetchone() is None else True

    def get_rows(self) -> list[tuple]:
        "Get all the rows of data from the table"
        if not self.__check_table_exists():
            print("No table exists")
            return [] # No table exist.
        sql_query = f"""
                SELECT * FROM {self.__table_name};
                """
        self.__db_c.execute(sql_query)
        string = self.__db_c.fetchall()
        return list(string)

    # data will be (workers:int, products:int, food:int)
    # def add_step(self, step:int, data = (0, 0, 0)):
    def add_step(self, data = (0, 0, 0)):
        "add step/day to the table, plug in new data"
        insert = f"""
            INSERT INTO {self.__table_name} (DATE, {', '.join(self.__table_columns)}) 
            VALUES (?, ?, ?, ?);
            """
        self.__db_c.execute(insert, (dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                                      *data)) #workers, products, food))
        self.__db.commit()

    def to_excel(self, filename: str):
        "Create an excel file from the tables data."
        if '.xlsx' not in filename:
            filename += '.xlsx'

        try:
            path = self.__getpath(self.__folder_name)
        except FileNotFoundError:
            path = ''
            print("Specific folder path not found, proceding anyway.")

        if not self.__check_table_exists():
            print("Table do not exist")
            return
        sql_quary = f"select date from {self.__table_name};"
        self.__db_c.execute(sql_quary)
        date = self.__db_c.fetchall()[0][0] #First value in the first tuple.

        try:
            wb = load_workbook(filename)
        except FileNotFoundError:
            wb = Workbook()
        # Skapar blad med datum och tid som namn på bladet
        ws = wb.create_sheet(date.replace(":","."))
        data = self.get_rows()

        for row in data:
            ws.append(row)
        # Sparar till samma filnamn
        if path == '':
            wb.save(filename)
        else:
            wb.save(path + "\\" + filename)


    def to_figure(self, filename:str):
        "Plot data in a figure and save the figure"
        file_path = ''
        try:
            file_path = self.__getpath(self.__folder_name) + '\\'
        except FileNotFoundError:
            file_path = ''


        if self.__check_table_exists():
            data = self.get_rows()
        else:
            if '.xlsx' not in filename:
                filename += '.xlsx'


            # Does not work with loading the excel file...
            try:
                df = pd.read_excel(file_path + filename)
            except FileNotFoundError:
                print("Did not find the file")
                return
            data = df
            print(data.head())
            print(data.columns)

        fig, ax = pyplot.subplots()

        arr_step = [row[0] for row in data ]

        # for resource in data:
        for i, label in enumerate(self.__table_columns):
            resource = [row[i + 2] for row in data]
            ax.plot(arr_step, resource, label = label)

        ax.set_xticks(arr_step)

        # Add some text for labels, title and custom x-axis tick labels, etc.
        ax.set_title('Resources in the simulation, by day')
        ax.set_ylabel('Amount')
        ax.set_xlabel('Day')
        ax.legend()

        fig.tight_layout()

        # Show the figure
        pyplot.show()
        # Export to file
        fig_filename = filename.rstrip('.xlsx') + '.png'
        fig.savefig(file_path + fig_filename)
        pyplot.close(fig)


#  Test this script.
if __name__ == "__main__":
    test = SimsimsAnalytics("Simsim_DB_test")
    test.drop_table()
    test.create_table()
    WORKERS= 2
    PRODUCTS=3
    FOOD=1
    test_data = (WORKERS, PRODUCTS, FOOD)
    test.add_step(data= test_data)
    WORKERS= WORKERS + 1
    PRODUCTS= PRODUCTS + 1
    FOOD= FOOD + 1
    test_data = (WORKERS, PRODUCTS, FOOD)
    test.add_step(data= test_data)
    WORKERS= WORKERS + 2
    PRODUCTS= PRODUCTS + 3
    FOOD= FOOD + 4
    test_data = (WORKERS, PRODUCTS, FOOD)
    test.add_step(data= test_data)

    print(test)
    print(test)

    testing = test.get_rows()

    if len(testing) == 10:
        test.drop_table()

    for test_row in testing:
        print(test_row)

    print(testing)

    test.to_excel("Testing_excel")

    test.to_figure('Testing_excel.xlsx')
